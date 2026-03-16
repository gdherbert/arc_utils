import arc_utils as au
import arcpy
import os


def write_field_info(fc):
    desc = arcpy.Describe(fc)
    root_path = desc.path
    fc_name = desc.name
    print('getting field info from %s' % fc_name)
    if fc_name.find(".") != -1:
        fc_name = fc_name.split(".")[0]
    a = open(os.path.join(root_path, fc_name + ".tsv"), "w")
    a.write(au.table.get_field_info_as_text(fc))  # use default tab
    a.flush()
    a.close()


def write_field_values(fc):
    import csv
    desc = arcpy.Describe(fc)
    root_path = desc.path
    print(root_path)
    fc_name = desc.name
    print('getting field values from %s' % fc_name)
    if fc_name.find(".") != -1:
        fc_name = fc_name.split(".")[0]
    field_value_csv = os.path.join(root_path, fc_name + "_field_values.csv")
    with open(field_value_csv, 'a+') as a:
        csvw = csv.writer(a, lineterminator='\n')
        fco = au.table.TableObj(fc)
        for field in fco.fields2:
            print('writing field values for %s' % field)
            values = [field]
            values.extend(fco.get_field_value_set(field))
            csvw.writerow(values)


def get_featureclass_info(workspace):
    ws = arcpy.env.workspace
    arcpy.env.workspace = workspace
    desc = arcpy.Describe(workspace)
    root_path = workspace
    if not desc.dataType == 'Folder':
        root_path = desc.catalogPath  # path
    for fc in arcpy.ListFeatureClasses():
        fc_path = os.path.join(arcpy.env.workspace, fc)
        # write field info
        write_field_info(fc_path)
        # write field values
        write_field_values(fc_path)
    arcpy.env.workspace = ws


def export_field_sets(layr_list, out_xlsx, ignore_fields=[], use_lyr_alias=True):
    """ Export unique field values for each layer in a list to an Excel file with one sheet per layer
    :param layr_list: list of layer objects or paths to layers
    :param out_xlsx: path to output Excel file
    :param ignore_fields: list of lowercase fields to ignore (e.g. ['objectid', 'shape'])
    :param use_lyr_alias: if True, use layer alias for sheet name;"""
	import os
	from openpyxl import Workbook

    ignore_fields = [v.lower() for v in ignore_fields]

	# Create workbook
	wb = Workbook()
	# Remove the default sheet that openpyxl creates
	default_sheet = wb.active
	wb.remove(default_sheet)

	# Loop through layers and create a sheet per layer

	for lyr in layr_list:
		tbl = au.table.TableObj(lyr)

		print(f'Processing name: {tbl.name}, alias: {str(lyr)}')

		# Excel sheet name must be <= 31 chars and cannot contain: : \ / ? * [ ]
		if use_lyr_alias:
			raw_name = str(lyr)
		else:
			raw_name = str(tbl.name) if hasattr(tbl, "name") else str(lyr)
		safe_name = "".join(ch for ch in raw_name if ch not in r':\/?*[]')
		safe_name = safe_name[:31] if len(safe_name) > 31 else safe_name
		if not safe_name:
			safe_name = "Sheet"

		ws = wb.create_sheet(title=safe_name)

		# Header row
		ws["A1"] = "Field"
		ws["B1"] = "Values"

		row_idx = 2

		for fld in tbl.fields2:
			if fld.lower() in ignore_fields:
                continue

            #get field alias if available
            if fld in tbl.field_dict and 'aliasName' in tbl.field_dict[fld] and tbl.field_dict[fld]['aliasName']:
                alias = tbl.field_dict[fld]['aliasName']
            else:
                alias = fld
			# Get unique values for the field
			vals = tbl.get_field_value_set(fld)

			# Convert to sorted, comma-separated string
			# Make sure everything is a string for consistency
			vals_str = ", ".join(sorted(str(v) for v in vals))

			# Write to Excel
            ws.cell(row=row_idx, column=1, value=alias)    # Column A: field alias/name
			try:
			    ws.cell(row=row_idx, column=2, value=vals_str) # Column B: CSV of unique values
            except Exception as e:
			    ws.cell(row=row_idx, column=2, value="VALUES ERROR") # Column B: error msg
                print(f"Error writing field '{alias}' values to Excel: {e}")
            row_idx += 1
                
	# Save workbook
	wb.save(out_xlsx)
	print(f"\nExcel file written to: {out_xlsx}")