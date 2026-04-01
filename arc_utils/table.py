# -*- coding: utf-8 -*-
"""utilities for working with tables, featureclasses and fields
"""
from __future__ import print_function, unicode_literals, absolute_import
import arcpy
import os
from .output import get_valid_output_path
from .output import output_msg


class TableObj(object):
    """ provide properties for working with a table/featureclass
    Usage: tbl = arc_utils.table.TableObj(path)
    :param
        path: a string representing an table/featureclass
    """
    def __init__(self, table_path):
        """ sets up reference to table
        adds properties and methods
        """
        self.path = table_path
        if isinstance(self.path, str):
            self.path = os.path.abspath(self.path)
        if not arcpy.Exists(self.path):
            raise RuntimeError("TableObj path is not found or invalid table/featureclass: {}".format(self.path))
        self.describe_obj = self._describe_object()
        self.name = self._get_fc_name()
        self.type = self._get_fc_type()
        self.field_dict = self._make_field_dict()
        self.fields = self._list_field_names()
        self.fields2 = self._list_field_names(required=False)
        self.fieldaliases = self._list_field_names(aliases=True)
        self.ignore_fields = ["objectid", "globalid","fid", "shape", "shape_area", "shape.area", "shape.starea()", "shape_length", "shape.len", "shape.stlength()"]
        

    def _describe_object(self):
        """ returns describe object"""
        return arcpy.Describe(self.path)

    def _get_fc_name(self):
        return self.describe_obj.baseName

    def _get_fc_type(self):
        if hasattr(self.describe_obj, 'shapeType'):
            return self.describe_obj.shapeType
        elif hasattr(self.describe_obj, 'dataType'):
            return self.describe_obj.dataType
        else:
            return 'Unknown'

    def _list_field_names(self, required=True, aliases=False):
        """Array of field names
        """
        if required:
            f_list = [field['name'] for field in self.field_dict.values()]
        else:
            f_list = [field['name'] for field in self.field_dict.values() if not field['required']]
        if aliases:
            # override the output to use aliases instead of field names
            f_list = [self.field_dict[field]['aliasName'] for field in f_list]
        return f_list

    def _make_field_dict(self):
        """Dictionary of fields containing
        all properties exposed by the arcpy.ListFields tool
        """
        field_dict = dict()
        for field in arcpy.ListFields(self.path):
            field_dict[field.name] = {
                "name": field.name,
                "baseName": field.baseName,
                "aliasName": field.aliasName,
                "type": field.type,
                "length": field.length,
                "required": field.required,
                "domain": field.domain,
                "defaultValue": field.defaultValue,
                "precision": field.precision,
                "scale": field.scale,
                "isNullable": field.isNullable,
                "editable": field.editable
            }
        return field_dict

    def get_field_info_as_text(self, sep="\t"):
        """ Create a delimeter separated output of a table's fields and their properties
            :param sep {String}
                Separator value to use. eg r"\t" for tab (default), "," for comma
            :return string of values with separator between
        """
        str_output = sep.join("{}\n".format(self.field_dict.keys()))
        for k in self.field_dict.keys():
            str_output += self.field_dict[k] + sep
        str_output += "\n"
        return str_output

    def get_max_field_value(self, field, lengthcomp=False):
        """Return the largest value (if numeric).
        lexicographic string comparison is used to determine largest value for strings by default.
            :param {String} field:
            name of the field to parse
            :param {Boolean} lengthcomp:
            If True will compare strings for length rather than lexicographically (ascii value of letters)
        """
        field_type = self.field_dict[field]['type']
        if field_type in ["Geometry"]:
            print("Cannot process Geometry field")
            return None
        elif field_type in ["String"]:
            result = ''
        else:
            result = 0
        with arcpy.da.SearchCursor(self.path, field) as values:
            for value in values:
                if value[0] is not None:
                    val = value[0]
                    if field_type in ["String"] and lengthcomp:
                        if len(val) > len(result):
                            result = val
                    else:
                        if val > result:
                            result = val
        return result

    def get_max_field_value_length(self, field):
        """Return the length of the maximum value in the field.
            :param: field {String}:
            name of the field to parse
        """
        length = 0
        with arcpy.da.SearchCursor(self.path, field) as values:
            for value in values:
                if value[0] is not None:
                    val = str(value[0])
                    if len(val) > length:
                        length = len(val)
        return length

    def get_field_value_set(self, field):
        """Return set of unique field values
            :param field {String}:
                name of the field to parse
            :return set of unique values. Null values are represented as 'NULL' string
           """
        if not arcpy.Exists(self.path):
            raise RuntimeError("TableObj path is not a table/featureclass: {}".format(self.path))
        try:
            value_set = set()  # set to hold unique values
            with arcpy.da.SearchCursor(self.path, field) as values:
                for value in values:
                    if value[0] is None:
                        value_set.add("NULL")
                    else:
                        value_set.add(value[0])
            return value_set

        except arcpy.ExecuteError:
            output_msg(arcpy.GetMessages(2))
        except Exception as e:
            output_msg(e.args[0])

    def export_fields_to_worksheet(self, worksheet, ignore_fields=None):
        """Write this table's field unique values to an openpyxl worksheet.
        if ignore_fields is not provided, will use the default ignore_fields property of the object
        if ignore_fields is provided, only those fields will be ignored
        :param worksheet: an openpyxl worksheet object to write to
        :param ignore_fields: a list of field names to ignore
        """
        if ignore_fields is None:
            ignore_fields = self.ignore_fields
        ignore_set = {v.lower() for v in ignore_fields}

        worksheet["A1"] = "Field"
        worksheet["B1"] = "Values"

        row = 2
        for field_name in self.fields2:
            if field_name.lower() in ignore_set:
                continue
            alias = self.field_dict.get(field_name, {}).get("aliasName") or field_name
            values = self.get_field_value_set(field_name)
            values_text = ", ".join(sorted(str(v) for v in values))
            worksheet.cell(row=row, column=1, value=alias)
            worksheet.cell(row=row, column=2, value=values_text)
            row += 1

        return row

    def get_multiple_field_value_set(self, fields, sep=':'):
        """return a set of unique field values for an input table
        and any number of fields (values will be concatenated with sep)
        :param fields {array of String values}:
            single field name or an array of field names (['Field1', 'Field2'])
        :param sep {String}:
            character to use as a separator (default = ':'
        """
        if not isinstance(fields, list):
            fieldslist = [fields]
        else:
            fieldslist = fields

        result = set()
        
        with arcpy.da.SearchCursor(self.path, fieldslist) as cursor:
            for row in cursor:
                parts = [('NULL' if v is None else str(v)) for v in row]
                result.add(sep.join(parts))

        return set(result)

    def find_duplicate_field_values(self, field, charset='ascii'):
        """Return set of unique field values
            :param field {String}:
                name of the field to parse
            :param: charset {String}:
                character set to use (default = 'ascii').
                Valid values are those in the Python documentation for string encode.
            :return set of values which are duplicated in the field (ignores Null values).
           """
        try:
            dup_set = set()  # set to hold duplicate values
            value_set = set()  # set to hold unique values
            with arcpy.da.SearchCursor(self.path, field) as values:
                for value in values:
                    if value[0] in value_set:
                        dup_set.add(value[0])
                    else:
                        value_set.add(value[0])
            return dup_set

        except arcpy.ExecuteError:
            output_msg(arcpy.GetMessages(2))
        except Exception as e:
            output_msg(e.args[0])

    def find_duplicate_field_values_as_df(path, fields):
        """Return set of unique field values
            :param path {String}:
                path to data
            :param field [{String}]:
                list of name(s) of fields to parse
            :return Pandas DataFrame of values which are duplicated in the field with count > 1 (ignores Null values).
            Use Pandas DataFrame <result>.to_csv() to export to file
        """
        if not isinstance(fields, list):
            fieldslist = [fields]
        else:
            fieldslist = fields
        
        import pandas
        data = arcpy.da.TableToNumPyArray(path, fieldslist, null_value='NULL')
        df = pandas.DataFrame(data)
        count = df.groupby(fields).size().reset_index().rename(columns={0:'count'})
        dups = count[count['count'] > 1]
        return dups
    
    def export_schema_to_csv(self, path):
        """Create a csv schema report of all fields in a featureclass,
        to the supplied path.
        """
        import datetime
        import os
        import csv
        # TODO convert to use csv
        start_time = datetime.datetime.today()
        start_date_string = start_time.strftime('%Y%m%d')
        default_env = arcpy.env.workspace
        fc = self.path
        delimiter = ','
        # nice to convert reported types to types accepted by add field tool
        type_conversions = {"String": "TEXT", "Float": "FLOAT", "Double": "DOUBLE", "SmallInteger": "SHORT",
                            "Integer": "LONG", "Date": "DATE", "Blob": "BLOB", "Raster": "RASTER", "GUID": "GUID",
                            "TRUE": "True", "FALSE": "False", "Geometry": "GEOMETRY", "OID": "OID"}

        output_msg("Processing: {}".format(self.path))

        try:
            fc_type = self.type

            report_dir = get_valid_output_path(path)
            out_file_name = self.name + "_Field_Report " + start_date_string + ".csv"
            out_file_path = os.path.join(report_dir, out_file_name)
            output_msg("Report file: {0}".format(out_file_path))
            with open(out_file_path, "w") as logFile:
                #logFile.write("Type,{}\n".format(fc_type))
                logFile.write(
                    "FieldName,FieldType,FieldPrecision,FieldScale,FieldLength,FieldAlias,isNullable,Required,"
                    "FieldDomain,DefaultValue,Editable,BaseName\n")

                for field in self.field_dict:
                    output_msg("Writing {}".format(field))
                    logFile.write("{0},{1},{2},{3},{4},{5},{6},{7},{8},{9},{10},{11}\n".format(
                        self.field_dict[field]['name'],
                        type_conversions[self.field_dict[field]['type']],
                        self.field_dict[field]['precision'],
                        self.field_dict[field]['scale'],
                        self.field_dict[field]['length'],
                        self.field_dict[field]['aliasName'],
                        self.field_dict[field]['isNullable'],
                        self.field_dict[field]['required'],
                        self.field_dict[field]['domain'],
                        self.field_dict[field]['defaultValue'],
                        self.field_dict[field]['editable'],
                        self.field_dict[field]['baseName'])
                    )

            return out_file_path
        except Exception as e:
            output_msg("error: " + str(e.args[0]))
            output_msg(arcpy.GetMessages())

    def compare_field_values_to_domain(self, field, gdb, domain_name):
        """compare field values with domain values
            return a named tuple (matched = values in domain,
            unmatched = values outside of domain
            NULL is skipped

            :param field {string}
                Field name
            :param gdb {string}
                Geodatabase path
            :param domain_name {string}
                Domain name in gdb
        """
        from collections import namedtuple
        nt = namedtuple('Result', 'match unmatched')
        field_values = self.get_field_value_set(field)
        domain_values = []
        domain_type = None
        field_in_domain = []
        field_out_domain = []
        domains = arcpy.da.ListDomains(gdb)
        for domain in domains:
            if domain.name == domain_name:
                if domain.domainType == 'CodedValue':
                    domain_type = 'CodedValue'
                    coded_values = domain.codedValues
                    for code, descr in coded_values.items():
                        domain_values.append(code)
                elif domain.domainType == 'Range':
                    domain_type = 'Range'
                    domain_values.append(domain.range[0])
                    domain_values.append(domain.range[1])
                break
        if domain_type == 'Range':
            # compare upper and lower bounds
            for fv in field_values:
                if fv == 'NULL':
                    continue
                elif int(fv) < domain_values[0] or int(fv) > domain_values[1]:
                    field_out_domain.append(fv)
                else:
                    field_in_domain.append(fv)
        else:
            for fv in field_values:
                if type(fv) == bytes: # Python 3 catch as domain values are strings
                    fv = fv.decode()
                if fv in domain_values:
                    field_in_domain.append(fv)
                else:
                    field_out_domain.append(fv)

        return nt(field_in_domain, field_out_domain)

    def pretty_print(self):
        """ pretty print a table's fields and their properties
        """
        def _print(l):
            print("".join(["{:>12}".format(i) for i in l]))

        atts = ['name', 'aliasName', 'type', 'baseName', 'domain',
                'editable', 'isNullable', 'length', 'precision',
                'required', 'scale', ]
        _print(atts)

        for f in arcpy.ListFields(self.path):
            _print(["{:>12}".format(getattr(f, i)) for i in atts])




def export_field_sets(fc_list, out_xlsx, ignore_fields=None, use_lyr_alias=True):
    """Export unique field values for each layer in a list to an Excel file with one sheet per layer.
    :param fc_list: list of feature class or table paths
    :param out_xlsx: path to output Excel file
    :param ignore_fields: list of lowercase fields to ignore (e.g. ['objectid', 'shape'])
    :param use_lyr_alias: if True, use layer alias for sheet name
    """
    from openpyxl import Workbook

    if ignore_fields is None:
        ignore_fields = ["objectid", "globalid"]
    ignore_fields = {v.lower() for v in ignore_fields}

    if not isinstance(fc_list, (list, tuple, set)):
        fc_list = [fc_list]

    wb = Workbook()
    # Remove the default sheet that openpyxl creates
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Loop through layers and create a sheet per layer
    for lyr in fc_list:
        tbl = TableObj(lyr)
        print(f'Processing name: {tbl.name}, alias: {str(lyr)}')

        # Excel sheet name must be <= 31 chars and cannot contain: : \ / ? * [ ]
        if use_lyr_alias:
            raw_name = str(lyr)
        else:
            raw_name = str(tbl.name) if hasattr(tbl, "name") else str(lyr)
        safe_name = "".join(ch for ch in raw_name if ch not in r':\/?*[]')
        safe_name = safe_name[:31] if len(safe_name) > 31 else safe_name
        if not safe_name:
            safe_name = "Sheet"

        ws = wb.create_sheet(title=safe_name)
        tbl.export_fields_to_worksheet(ws, ignore_fields=ignore_fields)
    # Save workbook
    wb.save(out_xlsx)
    print(f"\nExcel file written to: {out_xlsx}")


def compare_schema(fc1, fc2):
    """compare the schemas of two tables. Return an array of results.
    :param fc1 {String}:
            Path or reference to feature class or table.
    :param fc2 {String}:
            Path or reference to feature class or table.
    :return array of results (field not found, field same, etc)
    """
    result_arr= []
    fcobj1 = TableObj(fc1)
    fcobj2 = TableObj(fc2)
    field_dict1 = fcobj1.field_dict
    field_dict2 = fcobj2.field_dict
    for ifield in sorted(set(list(field_dict1.keys()) + list(field_dict2.keys()))):
        # check name for missing fields first
        if not ifield in field_dict1:
            the_result = "{0} not found in {1}".format(ifield, fc1)
            output_msg(the_result)
            result_arr.append(the_result)
        elif not ifield in field_dict2:
            the_result = "{0} not found in {1}".format(ifield, fc2)
            output_msg(the_result)
            result_arr.append(the_result)
        else:
            # string comparison of name, type and length
            if field_dict1[ifield] == field_dict2[ifield]:
                the_result = "{0} field same in both".format(ifield)
                output_msg(the_result)
                result_arr.append(the_result)
            else:
                the_result = "Field Mismatch {}: ({}, {}) ".format(
                    ifield, fc1, fc2)
                for k in field_dict1[ifield]:
                    if field_dict1[ifield][k] != field_dict2[ifield][k]:
                        the_result += '{}: ({}, {}) '.format(k, field_dict1[ifield][k], field_dict2[ifield][k])
                output_msg(the_result)
                result_arr.append(the_result)

    return result_arr


def import_schema_to_fc(csv_file, fc_name):
    """convert csv schema from report_fields_to_csv_schema
    to a featureclass"""
    # assume headers as per output from export_schema_to_csv
    # get fc type from csv file line 1
    # load the csv file from line 2 for field info
    # get field list
    # ignore OBJECTID, SHAPE
    fields_to_ignore =["OBJECTID", "GLOBALID","FID", "SHAPE", "SHAPE_AREA", "SHAPE.AREA", "SHAPE.STAREA()", "SHAPE_LENGTH", "SHAPE.LEN", "SHAPE.STLENGTH()"]
    # create fc from fc_name
    # add fields from field list
    # arcpy.AddField_management(in_table, field_name, field_type, {field_precision}, {field_scale}, {field_length}, {field_alias}, {field_is_nullable}, {field_is_required}, {field_domain})
    return None
